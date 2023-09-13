import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook

file = 'transactions.xlsx'
def process_workbook(fileName):
    wb = xl.load_workbook(fileName)
    sheet = wb['Sheet1']

    # Workbook().active #for create new file
    # cell = sheet.cell(3,2)
    # print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        print(row)
        cell = sheet.cell(row, 3)
        new_price = cell.value * 0.9
        print(new_price)
        new_price_cell = sheet.cell(row,5)

        new_price_cell.value = new_price


    # value = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)
    # chart = BarChart().add_data(value)
    # sheet.add_chart(chart, 'm2')

    wb.save('newfile1.xlsx')
# todo process_workbook(file)



# def my_xl():
#     wb = Workbook().active
#     # ws1 = wb.create_sheet("Mysheet")  # insert at the end (default)
#     ws1 = wb.create_sheet("sheet1", 0)
#     # ws2 = wb.create_sheet("sheet2")
#     # ws.title = "New Title"
#     # sheet1 = wb['Sheet1']
#     # sheet2 = wb['Sheet2']
#
#     # print(sheet1)
#     print(wb)
#     # print(ws1)

# my_xl()


print('eriufgyrukfgk3wf,')